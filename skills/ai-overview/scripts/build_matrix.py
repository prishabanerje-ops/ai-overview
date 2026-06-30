#!/usr/bin/env python3
"""
Build the AI Overview / ChatGPT citation-position analysis from one or more
scraped datasets (one per engine). Generic for ANY topic.

Produces an .xlsx with, FOR EACH engine:
  - "<Engine> positions" sheet: rows = prompts, columns = every domain cited in
    >= --min-prompts prompts (ranked by citation frequency). Cell = the domain's
    position in that answer's cited-sources list (1 = first cited), "NA" = not cited.
    The --brand domain is always shown and highlighted, even if below the threshold.
    Two summary rows at the bottom: "cited in" and "avg position".
  - "<Engine> text" sheet: the verbatim answer text + ordered citations per prompt.

Dataset JSON schema (written by the ai-overview-scraper agent), per engine:
{
  "topic": "Health Insurance",
  "engine": "google" | "chatgpt",
  "region": {"google": "google.ae", "gl": "ae", "hl": "en"},
  "date": "2026-06-24",
  "brand": "policybazaar.ae",
  "keywords": [
    {"q": "prompt", "present": true,  "text": "verbatim...", "hosts": ["a.com","b.ae", ...]},
    {"q": "prompt", "present": false, "text": "",            "hosts": []}
  ]
}
("present" = an answer/overview with sources was captured. Back-compat: the loader
also accepts the old "hasAIO" field, or infers presence from a non-empty hosts list.)

Usage (one or more engines):
  python3 build_matrix.py \
     --data "Google AIO=/path/google.json" "ChatGPT=/path/chatgpt.json" \
     --out report.xlsx [--brand policybazaar.ae] [--min-prompts 2]

A bare path with no "Label=" prefix is labelled "AI Overview".
"""
import argparse, json, re, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="D8D8D8")
BORDER = Border(THIN, THIN, THIN, THIN)


def norm(host: str) -> str:
    h = (host or "").strip().lower()
    for p in ("www.", "m.", "amp."):
        if h.startswith(p):
            h = h[len(p):]
    return h


def present(k) -> bool:
    if "present" in k:
        return bool(k["present"])
    if "hasAIO" in k:
        return bool(k["hasAIO"])
    return bool(k.get("hosts"))


def position(hosts, dom):
    for i, h in enumerate(hosts):
        if norm(h) == dom:
            return i + 1
    return "NA"


def safe_sheet(name: str) -> str:
    name = re.sub(r"[:\\/?*\[\]]", "-", name)
    return name[:31] if len(name) > 31 else name


def build_engine_sheets(wb, label, data, brand, min_prompts, first):
    KW = data.get("keywords", [])
    n = len(KW)
    cov = {}
    for k in KW:
        for d in {norm(h) for h in k.get("hosts", []) if norm(h)}:
            cov[d] = cov.get(d, 0) + 1
    cols = sorted([d for d, c in cov.items() if c >= min_prompts], key=lambda d: (-cov[d], d))
    if brand not in cols:
        cols.append(brand)

    ws = wb.active if first else wb.create_sheet()
    ws.title = safe_sheet(f"{label} positions")

    reg = data.get("region", {})
    regtxt = f"{reg.get('google','')} gl={reg.get('gl','')} hl={reg.get('hl','')}".strip()
    ws["A1"] = f"{label} — citation position by prompt  |  topic: {data.get('topic','')}  |  {regtxt}  |  {data.get('date','')}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = ("Value = position in the cited-sources list (1 = first cited). NA = not cited. "
                f"Columns = domains cited in ≥{min_prompts} prompts, by frequency. Tracked brand "
                f"({brand}) highlighted and always shown.")
    ws["A2"].font = Font(italic=True, size=9, color="666666")

    HROW = 4
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    brand_hdr = PatternFill("solid", fgColor="C55A11")
    hdr_font = Font(bold=True, color="FFFFFF", size=9)
    c = ws.cell(row=HROW, column=1, value="Prompt")
    c.fill = PatternFill("solid", fgColor="404040"); c.font = Font(bold=True, color="FFFFFF"); c.border = BORDER
    for j, d in enumerate(cols, start=2):
        cell = ws.cell(row=HROW, column=j, value=d)
        cell.font = hdr_font; cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="bottom", textRotation=90)
        cell.fill = brand_hdr if d == brand else hdr_fill

    brand_fill = PatternFill("solid", fgColor="FCE4D6")
    na_font = Font(color="C00000", size=9)
    cited_font = Font(color="1F7A1F", bold=True, size=9)
    counts = {d: 0 for d in cols}; possum = {d: 0 for d in cols}
    r = HROW + 1
    for k in KW:
        hosts = k.get("hosts", [])
        ws.cell(row=r, column=1, value=k.get("q", "")).border = BORDER
        for j, d in enumerate(cols, start=2):
            v = position(hosts, d)
            cell = ws.cell(row=r, column=j, value=v)
            cell.alignment = Alignment(horizontal="center"); cell.border = BORDER
            if d == brand:
                cell.fill = brand_fill
            if v == "NA":
                cell.font = na_font
            else:
                cell.font = cited_font; counts[d] += 1; possum[d] += v
        r += 1
    r += 1
    ws.cell(row=r, column=1, value=f"Cited in (of {n})").font = Font(bold=True)
    for j, d in enumerate(cols, start=2):
        ws.cell(row=r, column=j, value=f"{counts[d]} ({round(100*counts[d]/n) if n else 0}%)").font = Font(bold=True, size=9)
        ws.cell(row=r, column=j).alignment = Alignment(horizontal="center")
    r += 1
    ws.cell(row=r, column=1, value="Avg position when cited").font = Font(bold=True)
    for j, d in enumerate(cols, start=2):
        avg = round(possum[d] / counts[d], 1) if counts[d] else "NA"
        ws.cell(row=r, column=j, value=avg).font = Font(bold=True, size=9)
        ws.cell(row=r, column=j).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 52
    for j in range(2, len(cols) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 6.5
    ws.row_dimensions[HROW].height = 130
    ws.freeze_panes = "B5"

    # text sheet
    ws2 = wb.create_sheet(safe_sheet(f"{label} text"))
    for j, head in enumerate(["Prompt", "Has answer?", "# citations", "Verbatim text", "Cited domains (in order)"], 1):
        cc = ws2.cell(row=1, column=j, value=head)
        cc.font = Font(bold=True, color="FFFFFF"); cc.fill = hdr_fill
    rr = 2
    for k in KW:
        ws2.cell(row=rr, column=1, value=k.get("q", ""))
        ws2.cell(row=rr, column=2, value="Yes" if present(k) else "No")
        ws2.cell(row=rr, column=3, value=len(k.get("hosts", [])))
        tcell = ws2.cell(row=rr, column=4, value=k.get("text", ""))
        tcell.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.cell(row=rr, column=5, value=", ".join(norm(h) for h in k.get("hosts", [])))
        rr += 1
    ws2.column_dimensions["A"].width = 46
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 11
    ws2.column_dimensions["D"].width = 110
    ws2.column_dimensions["E"].width = 60
    ws2.freeze_panes = "A2"

    # leaderboard for stdout
    rank = sorted(cols, key=lambda d: (-counts[d], d))
    lb = [(d, counts[d], round(possum[d]/counts[d], 1) if counts[d] else "NA") for d in rank]
    answered = sum(1 for k in KW if present(k))
    return {"label": label, "n": n, "answered": answered, "leaderboard": lb,
            "brand": {"cited": counts.get(brand, 0),
                      "avg": round(possum[brand]/counts[brand], 1) if counts.get(brand) else "NA"}}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", required=True, nargs="+", help='one or more "Label=path.json" (or bare path)')
    ap.add_argument("--out", required=True)
    ap.add_argument("--brand", default="policybazaar.ae")
    ap.add_argument("--min-prompts", type=int, default=2)
    args = ap.parse_args()
    brand = norm(args.brand)

    engines = []
    for tok in args.data:
        if "=" in tok:
            label, path = tok.split("=", 1)
        else:
            label, path = "AI Overview", tok
        engines.append((label.strip(), path.strip()))

    wb = Workbook()
    summaries = []
    for i, (label, path) in enumerate(engines):
        try:
            data = json.load(open(path))
        except Exception as e:
            sys.exit(f"Could not read dataset for '{label}' at {path}: {e}")
        if not data.get("keywords"):
            sys.exit(f"Dataset for '{label}' has no keywords.")
        summaries.append(build_engine_sheets(wb, label, data, brand, args.min_prompts, first=(i == 0)))
    wb.save(args.out)

    print(f"Wrote {args.out}")
    for s in summaries:
        print(f"\n=== {s['label']} ===  prompts: {s['n']}  |  with answer+sources: {s['answered']}")
        print("Rank  Domain                              Cited        Avg pos")
        for i, (d, c, avg) in enumerate(s["leaderboard"], 1):
            tag = "  <- brand" if d == brand else ""
            print("%3d   %-34s %2d/%d (%3d%%)   %s%s" % (i, d, c, s["n"], round(100*c/s["n"]) if s["n"] else 0, avg, tag))
    # combined brand line
    print("\nBrand visibility across engines:")
    for s in summaries:
        print(f"  {s['label']:<14} {brand}: cited {s['brand']['cited']}/{s['n']}, avg pos {s['brand']['avg']}")


if __name__ == "__main__":
    main()
